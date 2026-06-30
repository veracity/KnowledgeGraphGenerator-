import csv
import hashlib
import json
import os

from data.table import Table, unique_columns

class Data:
    __path: str
    __name: str
    __type: str
    __df: Table
    __loaded: bool
    __error: bool

    def __init__(self, path: str, sheet: str = None) -> None:
        self.__path = path
        base = os.path.basename(path)
        name, ext = os.path.splitext(base)
        self.__type = ext.lstrip(".").lower()
        self.__sheet = sheet
        self.__name = f"{name} - {sheet}" if sheet is not None else name
        self.__loaded = False
        self.__error = False
        self.__checksum = None
        pass

    @property
    def path(self) -> str:
        return self.__path

    @property
    def name(self) -> str:
        return self.__name

    @property
    def id(self) -> str:
        return self.__name

    @property
    def type(self) -> str:
        return self.__type

    @property
    def sheet(self) -> str:
        return self.__sheet

    @property
    def checksum(self) -> str:
        if self.__checksum is None:
            try:
                h = hashlib.sha256()
                with open(self.__path, "rb") as f:
                    for chunk in iter(lambda: f.read(8192), b""):
                        h.update(chunk)
                self.__checksum = h.hexdigest()
            except OSError:
                self.__checksum = ""
        return self.__checksum

    @property
    def df(self) -> Table:
        return self.__df

    @property
    def columns(self) -> list:
        if not self.__loaded:
            self.loadData()
        df = getattr(self, "_Data__df", None)
        return list(df.columns) if df is not None else []

    @property
    def loaded(self) -> bool:
        return self.__loaded

    @property
    def error(self) -> bool:
        return self.__error

    def __str__(self) -> str:
        return f"path: {self.__path}, name: {self.__name}, type: {self.__type}"

    def __repr__(self) -> str:
        return self.__str__()

    def __eq__(self, __o: object) -> bool:
        if not isinstance(__o, Data): return False
        if (self.__path != __o.path): return False
        if (self.__sheet != __o.sheet): return False
        return True
    
    def __hash__(self) -> int:
        return hash(f"{self.__path}{self.__sheet}")

    def loadData(self) -> None:
        self.__error = False
        try:
            if self.__type == "csv":
                self.__df = _read_csv(self.__path)
            elif self.__type == "xlsx":
                self.__df = _read_excel(self.__path, self.__sheet)
            elif self.__type == "json":
                self.__df = _read_json(self.__path)
        except Exception as e:
            print(e)
            print("could not load")
            self.__error = True
        self.__loaded = not self.__error
        return


def _read_csv_with(path: str, delimiter: str) -> list:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f, delimiter=delimiter))


def _read_csv(path: str) -> Table:
    rows = _read_csv_with(path, ";")
    if rows and len(rows[0]) == 1:
        alt = _read_csv_with(path, ",")
        if alt and len(alt[0]) > 1:
            rows = alt
    header = unique_columns(rows[0]) if rows else []
    data = {column: [] for column in header}
    for row in rows[1:]:
        for index, column in enumerate(header):
            data[column].append(row[index] if index < len(row) else "")
    return Table(header, data, types={column: "string" for column in header})


def _read_excel(path: str, sheet: str) -> Table:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = wb[sheet] if sheet is not None else wb[wb.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        first = next(rows, None)
        header = unique_columns(first) if first is not None else []
        data = {column: [] for column in header}
        for row in rows:
            for index, column in enumerate(header):
                data[column].append(row[index] if index < len(row) else None)
    finally:
        wb.close()
    return Table(header, data)


def _read_json(path: str) -> Table:
    with open(path, encoding="utf-8") as f:
        obj = json.load(f)

    if isinstance(obj, list):
        columns = []
        for record in obj:
            for key in record.keys():
                if key not in columns:
                    columns.append(key)
        data = {column: [record.get(column) for record in obj] for column in columns}
        return Table(columns, data)

    if isinstance(obj, dict):
        columns = list(obj.keys())
        data = {}
        for column in columns:
            value = obj[column]
            if isinstance(value, dict):
                data[column] = list(value.values())
            elif isinstance(value, list):
                data[column] = list(value)
            else:
                data[column] = [value]
        return Table(columns, data)

    raise ValueError("Unsupported JSON structure")


def excel_sheet_names(path: str) -> list:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return []